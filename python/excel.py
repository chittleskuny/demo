#!/usr/bin/python
# -*- coding:utf-8 -*-


import os
import openpyxl
import itertools
import xlrd, xlwt


class DemoExcel(object):
    cell_width = { 'xls': 256*11, 'xlsx': 8.38 }

    def __init__(self, workbook_name=None):
        self.workbook = None
        self.origin = 0
        self.mode = 'r'

        self.open_workbook(workbook_name)

    ### workbook ###

    def new_workbook(self, workbook_name=None):
        if workbook_name.endswith('.xls'):
            self.workbook = xlwt.Workbook()
            self.extension = '.xls'
            self.origin = 1
        elif workbook_name.endswith('.xlsx'):
            self.workbook = openpyxl.Workbook()
            self.extension = '.xlsx'
            self.origin = 0
        else:
            self.workbook = None
        self.name = workbook_name
        self.mode = 'w'
        return self.workbook

    def open_workbook(self, workbook_name=None):
        if not workbook_name or not os.path.isfile(workbook_name):
            self.new_workbook(workbook_name)
        else:
            if workbook_name.endswith('.xls'):
                self.workbook = xlrd.open_workbook(filename=workbook_name)
                self.extension = '.xls'
            elif workbook_name.endswith('.xlsx'):
                self.workbook = openpyxl.load_workbook(filename=workbook_name)
                self.extension = '.xlsx'
            else:
                pass
            self.name = workbook_name
            self.mode = 'r'
        return self.workbook

    def save_workbook(self, workbook_name=None):
        if self.mode == 'r':
            return None
        elif self.mode == 'w':
            self.workbook.save(workbook_name if workbook_name else self.name)
        else:
            pass
        return self.workbook

    ### worksheets ###

    def get_worksheets(self):
        if self.mode == 'r':
            if self.extension == '.xls':
                sheets = self.workbook.sheets()
            elif self.extension == '.xlsx':
                sheets = None
            else:
                sheets = None
        elif self.mode == 'w':
            sheets = None
        else:
            sheets = None
        return sheets

    def get_worksheet_names(self):
        if self.mode == 'r':
            if self.extension == '.xls':
                sheet_names = self.workbook.sheet_names()
            elif self.extension == '.xlsx':
                sheet_names = None
            else:
                sheet_names = None
        elif self.mode == 'w':
            sheet_names = None
        else:
            sheet_names = None
        return sheet_names

    ### worksheet ###

    def new_worksheet(self, worksheet_name):
        if self.mode == 'r':
            return None
        elif self.mode == 'w':
            if self.extension == '.xls':
                worksheet = self.workbook.add_sheet(worksheet_name)
            elif self.extension == '.xlsx':
                worksheet = self.workbook.create_sheet(worksheet_name)
            else:
                worksheet = None
        else:
            worksheet = None
        return worksheet

    def open_worksheet(self, worksheet_name=None, worksheet_index=None):
        if worksheet_name and worksheet_index:
            return None
        if self.mode == 'r':
            if self.extension == '.xls':
                if worksheet_name:
                    worksheet = self.workbook.sheet_by_name(worksheet_name)
                elif worksheet_index:
                    worksheet = self.workbook.sheet_by_index(worksheet_index)
                else:
                    worksheet = None
            elif self.extension == '.xlsx':
                worksheet = None
            else:
                worksheet = None
        elif self.mode == 'w':
            if self.extension == '.xls':
                if worksheet_name:
                    try:
                        for idx in itertools.count():
                            worksheet_temp = self.workbook.get_sheet(idx)
                            if worksheet_temp.name == worksheet_name:
                                worksheet = worksheet_temp
                                break
                    except IndexError:
                        worksheet = self.new_worksheet(worksheet_name)
                elif worksheet_index:
                    worksheet = self.workbook.get_sheet(worksheet_index)
                else:
                    worksheet = None
            elif self.extension == '.xlsx':
                if worksheet_name:
                    if worksheet_name in self.workbook:
                        worksheet = self.workbook[worksheet_name]
                    else:
                        worksheet = self.new_worksheet(worksheet_name)
                elif worksheet_index:
                    pass
                else:
                    worksheet = None
            else:
                worksheet = None
        else:
            worksheet = None
        return worksheet

    ### origin ###

    def get_worksheet_origin(self):
        return self.origin

    ### ranges ###

    def get_worksheet_ranges(self):
        if self.mode == 'r':
            if self.extension == '.xls':
                sheet_ranges = None
            elif self.extension == '.xlsx':
                sheet_ranges = self.workbook.get_named_ranges()
            else:
                sheet_ranges = None
        elif self.mode == 'w':
            sheet_ranges = None
        else:
            sheet_ranges = None
        return sheet_ranges

    def get_number_of_columns(self, worksheet):
        if self.mode == 'r':
            if self.extension == '.xls':
                number_of_columns = worksheet.ncolumns()
            elif self.extension == '.xlsx':
                number_of_columns = worksheet.get_highest_column()
            else:
                number_of_columns = None
        elif self.mode == 'w':
            number_of_columns = None
        else:
            number_of_columns = None
        return number_of_columns

    def get_number_of_rows(self, worksheet):
        if self.mode == 'r':
            if self.extension == '.xls':
                number_of_rows = worksheet.nrows()
            elif self.extension == '.xlsx':
                number_of_rows = worksheet.get_highest_row()
            else:
                number_of_rows = None
        elif self.mode == 'w':
            number_of_rows = None
        else:
            number_of_rows = None
        return number_of_rows

    ### columns & rows ###

    def get_column_values(self, worksheet, column_index):
        if self.mode == 'r':
            if self.extension == '.xls':
                column_values = self.workbook.column_values(column_index)
            elif self.extension == '.xlsx':
                column_values = None
            else:
                column_values = None
        elif self.mode == 'w':
            column_values = None
        else:
            column_values = None
        return column_values

    def get_row_values(self, worksheet, row_index):
        if self.mode == 'r':
            if self.extension == '.xls':
                row_values = self.workbook.row_values(row_index)
            elif self.extension == '.xlsx':
                row_values = None
            else:
                row_values = None
        elif self.mode == 'w':
            row_values = None
        else:
            row_values = None
        return row_values

    ### column & row ###

    def set_column_width(self, worksheet, column=None, column_index=None, multiple=1):
        if column and not column_index:
            column_index = self.get_column_index_from_string(column)
        if not column_index:
            return False

        if self.mode == 'r':
            return False
        elif self.mode == 'w':
            if self.extension == '.xls':
                worksheet.col(column_index).width = int(self.cell_width['xls']*multiple)
            elif self.extension == '.xlsx':
                worksheet.column_dimensions[column].width = self.cell_width['xlsx']*multiple
            else:
                return False
        else:
            return False
        return True

    ### cell ###

    def get_cell(self, worksheet, column=None, row=None, column_index=None, row_index=None):
        if not column_index:
            column_index = self.get_column_index_from_string(column)
        if not row_index:
            row_index = row

        if self.mode == 'r':
            if self.extension == '.xls':
                cell = worksheet.cell(row_index, column_index)
            elif self.extension == '.xlsx':
                cell = worksheet.cell(row_index, column_index)
            else:
                cell = None
        elif self.mode == 'w':
            cell = None
        else:
            cell = None
        return cell

    def get_cell_value(self, worksheet, column=None, row=None, column_index=None, row_index=None):
        if not column_index:
            column_index = self.get_column_index_from_string(column)
        if not row_index:
            row_index = row

        if self.mode == 'r':
            if self.extension == '.xls':
                cell_value = worksheet.cell_value(row_index, column_index)
            elif self.extension == '.xlsx':
                cell_value = worksheet.cell(row_index, column_index).value
            else:
                cell_value = None
        elif self.mode == 'w':
            cell_value = None
        else:
            cell_value = None
        return cell_value

    def set_cell(self, worksheet, column=None, row=None, column_index=None, row_index=None, merge_column_vector=0, row_merge_vector=0, cell_value=None, formula=None, style=None):
        if not column_index:
            column_index = self.get_column_index_from_string(column)
        if not row_index:
            row_index = row

        if self.mode == 'r':
            return False
        elif self.mode == 'w':
            if self.extension == '.xls':
                if merge_column_vector == 0 and row_merge_vector == 0:
                    if not style:
                        worksheet.write(row_index, column_index, label=cell_value)
                    else:
                        worksheet.write(row_index, column_index, label=cell_value, style=style)
                else:
                    if not style:
                        worksheet.write_merge(row_index, row_index + row_merge_vector, column_index, column_index + merge_column_vector, label=cell_value)
                    else:
                        worksheet.write_merge(row_index, row_index + row_merge_vector, column_index, column_index + merge_column_vector, label=cell_value, style=style)
            elif self.extension == '.xlsx':
                worksheet.cell(row_index, column_index).value = cell_value
            else:
                pass
        else:
            pass
        return True

    ### formula ###

    def get_formula(self, formula_string):
        if self.mode == 'r':
            formula = None
        elif self.mode == 'w':
            if self.extension == '.xls':
                formula = xlwt.Formula(formula_string)
            elif self.extension == '.xlsx':
                formula = None
            else:
                formula = None
        else:
            formula = None
        return formula

    ### style ###

    def get_style(self, xf_string):
        if self.mode == 'r':
            formula = None
        elif self.mode == 'w':
            if self.extension == '.xls':
                style = xlwt.Style.easyxf(xf_string)
            elif self.extension == '.xlsx':
                style = None
            else:
                style = None
        else:
            formula = None
        return style

    ### utils ###

    def get_coordinate_from_string(self, string):
        coordinate = openpyxl.utils.cell.coordinate_from_string(string)
        return coordinate

    def get_coordinate_from_coordinate_index(self, column_index, row_index):
        wso = self.get_worksheet_origin(self.workbook)
        column_letter = openpyxl.utils.cell.get_column_letter(column_index + wso)
        row_index = row_index + self.origin
        return column_letter + str(row_index)

    def get_column_index_from_string(self, string):
        (column_string, row_string) = openpyxl.utils.cell.coordinate_from_string(string)
        column_index = openpyxl.utils.cell.column_index_from_string(column_string) + self.origin
        return column_index

    def get_coordinate_index_from_string(self, string):
        wso = self.get_worksheet_origin(self.workbook)
        (column_string, row_string) = openpyxl.utils.cell.coordinate_from_string(string)
        column_index = openpyxl.utils.cell.column_index_from_string(column_string)
        row_index = int(row_string)
        coordinate_index = (column_index - wso, row_index - wso)
        return coordinate_index
